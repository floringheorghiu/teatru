#!/usr/bin/env python3
"""
teatre.py — staged harvester for Bucharest theatre programmes.

  Stage 0  preflight   — prove the network works BEFORE diagnosing anything
  Stage 1  discovery   — surse.json, edited by hand, persisted across runs
  Stage 2  harvest     — fetch + snapshot raw HTML to disk (the evidence)
  Stage 2b diagnose    — report what each snapshot actually contains
  Stage 3  extract     — one model call per page, strict JSON out
  Stage 4  verify      — mechanical: every row's quoted source string must
                         literally occur in the snapshot, else the row is dropped
  Stage 5  render      — always emit the event list (.xlsx, .csv fallback),
                         with a status per institution

Stage 3 is the only stage that needs a model. Everything else is deterministic
and therefore cannot be skipped, faked, or "claimed complete".

CONCURRENCY
    Network stages fan out with --concurrency (default 4): curl is a subprocess,
    so threads genuinely overlap, and browser-strategy pages share ONE Chromium
    (launch+teardown costs seconds each). Model calls fan out with
    --extract-concurrency (default 2 — API tiers limit tokens/minute, not
    requests, so 2x parallel is the safe free-tier speedup).

PERIODS
    --period current-month      the calendar month containing today
    --period next-month
    --period 2026-09            a named month
    --period 2026-09-12         a single day
    --period 2026-09-12..2026-09-20
    --date   2026-09-12         alias for a single day

Sites whose calendar takes a month parameter (ONB: ?luna=&anul=) are fetched
once per month in the range; the rest are fetched once and filtered by date.

Usage:
    python3 teatre.py preflight
    python3 teatre.py harvest  --period current-month
    python3 teatre.py diagnose --period current-month
    python3 teatre.py extract  --period current-month --backend ollama
    python3 teatre.py verify   --period current-month
    python3 teatre.py render   --period current-month
    python3 teatre.py all      --period current-month --backend ollama
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import json
import os
import pathlib
import re
import socket
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

TIMINGS: list[tuple[str, float, str]] = []


def fmt_dur(sec: float) -> str:
    if sec < 1:
        return f"{sec * 1000:.0f}ms"
    if sec < 60:
        return f"{sec:.1f}s"
    return f"{int(sec // 60)}m {sec % 60:04.1f}s"


def stage_done(title: str, t0: float, summary: str) -> None:
    """Every stage reports what it achieved and how long it took."""
    elapsed = time.monotonic() - t0
    TIMINGS.append((title, elapsed, summary))
    print(f"\n  ⏱  {title}: {summary}  [{fmt_dur(elapsed)}]")


def print_rollup() -> None:
    if len(TIMINGS) < 2:
        return
    total = sum(t for _, t, _ in TIMINGS)
    print("\n" + "=" * 72)
    print(f"  RUN SUMMARY — total {fmt_dur(total)}")
    print("=" * 72)
    for name, sec, summary in TIMINGS:
        bar = "█" * max(1, round(24 * sec / total))
        print(f"  {name:<10} {fmt_dur(sec):>9}  {bar:<24} {summary}")
    print("=" * 72)

ROOT = pathlib.Path(__file__).resolve().parent
SOURCES = ROOT / "surse.json"
SNAPDIR = ROOT / "snapshots"
OUTDIR = ROOT / "rezultate"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

RO_MONTHS = {
    1: ["ianuarie", "ian"], 2: ["februarie", "feb"], 3: ["martie", "mar"],
    4: ["aprilie", "apr"], 5: ["mai"], 6: ["iunie", "iun"],
    7: ["iulie", "iul"], 8: ["august", "aug"], 9: ["septembrie", "sept", "sep"],
    10: ["octombrie", "oct"], 11: ["noiembrie", "noi", "nov"],
    12: ["decembrie", "dec"],
}

# curl exit codes worth naming explicitly, so nobody ever again reads a DNS
# failure as "this site requires JavaScript".
CURL_ERRORS = {
    3: "malformed URL",
    5: "could not resolve proxy",
    6: "COULD NOT RESOLVE HOST — this is DNS. Not JavaScript, not the site.",
    7: "failed to connect to host — network/firewall, not the page content",
    28: "timeout",
    35: "TLS handshake failure",
    56: "connection reset / proxy refused CONNECT",
    60: "TLS certificate problem",
}


# ================================================================= patterns
# Every regex that runs more than once is compiled exactly once, here. Pages
# are 1-2 MB, and several hot paths used to rebuild pattern strings on every
# call — including a fresh pattern per Romanian month name inside
# period_signals, and full-page re-scans in the deterministic extractors.

TAG_STRIP_RE = re.compile(r"<[^>]+>")
WS_RE = re.compile(r"\s+")
WS_LINE_RE = re.compile(r"[ \t]+")
COMMENT_RE = re.compile(r"<!--.*?-->", re.S)
FENCE_RE = re.compile(r"^\s*```(?:json)?|```\s*$", re.M)

# condense() — the payload builder
NOISE_BLOCK_RE = re.compile(r"<(script|style|svg|noscript|iframe)\b.*?</\1>", re.S | re.I)
DATA_ATTR_TAG_RE = re.compile(r'<[a-z][^>]*?\bdata-[\w-]*(?:date|time|day|venue)[\w-]*="[^"]*"[^>]*>', re.I)
DATA_ATTR_VAL_RE = re.compile(r'\bdata-[\w-]*(?:date|time|day|venue)[\w-]*="([^"]{2,60})"', re.I)
ANCHOR_RE = re.compile(r'<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>', re.S | re.I)
NBSP_RE = re.compile(r"&nbsp;?")
BLOCK_BREAK_RE = re.compile(r"<(p|div|li|tr|section|article|table|ul|ol|h[1-6])\b[^>]*>", re.I)
TD_BREAK_RE = re.compile(r"<(td|th)\b[^>]*>", re.I)

# canary / diagnose — svg & style are coordinate/boilerplate noise; <script> is
# kept on purpose (Bulandra's whole programme lives in one)
SVG_STYLE_RE = re.compile(r"<(svg|style)\b.*?</\1>", re.S | re.I)
SCRIPT_STYLE_SVG_RE = re.compile(r"<(script|style|svg)\b.*?</\1>", re.S | re.I)

# period_signals — in-period date finders, all three formats
ISO_DATE_RE = re.compile(r"(?<!\d)(\d{4})-(\d{1,2})-(\d{1,2})(?!\d)")
DOT_DATE_RE = re.compile(r"(?<![\d.\-/])(\d{1,2})[./](\d{1,2})(?![\d.\-/:])")
MONTH_NAME_DATE_RE = {name: re.compile(rf"(?<!\d)(\d{{1,2}})\s*{name}\b", re.I)
                      for names in RO_MONTHS.values() for name in names}

# salvage_script_schedules — ISO stamps inside plain <script> payloads
SCRIPT_BODY_RE = re.compile(r"<script\b[^>]*>(.*?)</script>", re.S | re.I)
STAMP_RE = re.compile(r"\d{4}-\d{2}-\d{2}(?:T\d{2}:\d{2})?")
UNICODE_ESC_RE = re.compile(r"\\u([0-9a-fA-F]{4})")

# extract_structured — typed JSON payloads in <script> tags
SCRIPT_TAG_RE = re.compile(r"<script\b([^>]*)>(.*?)</script>", re.S | re.I)

# deterministic extractors
DATA_FULLDATE_RE = re.compile(r'data-fulldate="([^"]+)"', re.I)
DATA_VENUE_RE = re.compile(r'data-venue="([^"]+)"', re.I)
SPECTACOL_HREF_RE = re.compile(r'href="([^"]*/spectacol/[^"]*)"', re.I)
SPECTACOL_TITLE_RE = re.compile(r'/spectacol/[^"]*"[^>]*>([^<]{2,80})</a>', re.I)
HOUR_TEXT_RE = re.compile(r">\s*(\d{1,2}:\d{2})\s*<")
UPPER_SPACE_RE = re.compile(r"(?<!^)(?=[A-Z])")
INLINE_START_RE = re.compile(r'"start"\s*:\s*"([^"]+)"')
PERMALINK_RE = re.compile(r'"permalink"\s*:\s*"([^"]+)"')
SUBJECT_RE = re.compile(r"subject=([^&\"]+)")
SALA_RE = re.compile(r"Sala\s+[A-ZȘŢȚ][\w\-șşțţăâî]+(?:\s+[A-ZȘŢȚ][\w\-șşțţăâî]+)?")
SLUG_NUMS_RE = re.compile(r"-\d+(-\d+)*$")

# FullCalendar (TNB) — events are positioned by column under thead data-dates
DATA_DATE_RE = re.compile(r'data-date="(\d{4}-\d{2}-\d{2})"')
TR_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.S)
TD_SPLIT_RE = re.compile(r"(?=<td\b)")
COLSPAN_RE = re.compile(r'colspan="(\d+)"')
ROWSPAN_RE = re.compile(r'rowspan="(\d+)"')
FC_EVENT_RE = re.compile(r"<a class=\"fc-day-grid-event.*?</a>\s*</div>", re.S)
FC_TITLE_RE = re.compile(r"<h3[^>]*>(.*?)</h3>", re.S)
FC_HOUR_RE = re.compile(r'class="hour">\s*Ora:\s*([\d:]+)')
FC_LOC_RE = re.compile(r'class="location"[^>]*>(.*?)</span>', re.S)
FC_LINK_RE = re.compile(r'<a href="([^"]+)"')

# discover
CANONICAL_RE = re.compile(r'<link[^>]+canonical[^>]+href="(https?://[^/"]+)')
HREF_RE = re.compile(r'href="([^"]+)"')
ASSET_EXT_RE = re.compile(r"\.(jpg|jpeg|png|pdf|css|js)$", re.I)
QUERY_HASH_RE = re.compile(r"[?#].*$")
SCHEME_RE = re.compile(r"^https?://(www\.)?")
WP_NOISE_RE = re.compile(r"/wp-json/|/feed/?$|oembed|[?&]format=", re.I)
SCHEME_HOST_RE = re.compile(r"^https?://[^/]+")


# ================================================================= period
class Period:
    """An inclusive date range, plus the snapshot-directory label it owns."""

    def __init__(self, start: dt.date, end: dt.date, label: str, kind: str):
        if end < start:
            raise ValueError("period end precedes start")
        self.start, self.end, self.label, self.kind = start, end, label, kind

    @property
    def single_day(self) -> bool:
        return self.start == self.end

    def contains(self, d: dt.date) -> bool:
        return self.start <= d <= self.end

    def months(self) -> list[tuple[int, int]]:
        """Every (year, month) the range touches, in order."""
        out, y, m = [], self.start.year, self.start.month
        while (y, m) <= (self.end.year, self.end.month):
            out.append((y, m))
            y, m = (y + 1, 1) if m == 12 else (y, m + 1)
        return out

    def human(self) -> str:
        if self.single_day:
            return self.start.strftime("%d.%m.%Y")
        return f"{self.start.strftime('%d.%m.%Y')} – {self.end.strftime('%d.%m.%Y')}"

    def __str__(self) -> str:
        return self.label


def month_period(year: int, month: int) -> Period:
    last = calendar.monthrange(year, month)[1]
    return Period(dt.date(year, month, 1), dt.date(year, month, last),
                  f"{year}-{month:02d}", "month")


def parse_period(spec: str, today: dt.date | None = None) -> Period:
    """Accept CURRENT_MONTH, next-month, YYYY-MM, YYYY-MM-DD, or A..B."""
    today = today or dt.date.today()
    s = spec.strip().lower().replace("_", "-")

    if s in ("current-month", "this-month", "luna-curenta", "luna-curentă"):
        return month_period(today.year, today.month)
    if s in ("next-month", "luna-viitoare"):
        y, m = (today.year + 1, 1) if today.month == 12 else (today.year, today.month + 1)
        return month_period(y, m)
    if s in ("today", "azi"):
        return Period(today, today, today.isoformat(), "day")

    if ".." in s:
        a, _, b = s.partition("..")
        start, end = dt.date.fromisoformat(a.strip()), dt.date.fromisoformat(b.strip())
        return Period(start, end, f"{start.isoformat()}_{end.isoformat()}", "range")

    if re.fullmatch(r"\d{4}-\d{2}", s):
        return month_period(int(s[:4]), int(s[5:]))

    d = dt.date.fromisoformat(s)
    return Period(d, d, d.isoformat(), "day")


# ================================================================= fetching
def load_sources(include_disabled: bool = False) -> list[dict]:
    """Active institutions by default.

    An institution disabled in surse.json is skipped by every working stage but
    still listed by `render`, marked as excluded with its reason. Deleting the
    entry instead would make a 12-venue brief silently become an 11-venue one —
    the same "did we actually check it?" ambiguity this pipeline exists to kill.
    """
    if not SOURCES.exists():
        sys.exit(f"missing {SOURCES} — run discovery first (stage 1)")
    rows = json.loads(SOURCES.read_text(encoding="utf-8"))["institutions"]
    return rows if include_disabled else [i for i in rows if i.get("enabled", True)]


def host_of(url: str) -> str:
    return re.sub(r"^https?://", "", url).split("/")[0]


def is_month_parameterised(url: str) -> bool:
    return any(t in url for t in ("{MM}", "{M}", "{YYYY}"))


def fmt_url(url: str, year: int, month: int) -> str:
    """Fill month/year placeholders.

    {MM} zero-padded (09), {M} bare (9), {YYYY} year. Both month forms are needed:
    ONB wants ?luna=09 while TNB wants ?month=9, and sending the padded form to a
    site expecting the bare one silently returns an empty calendar rather than an
    error — which this pipeline would then report as "venue has no schedule".
    """
    return url.format(MM=f"{month:02d}", M=str(month), YYYY=year)


def curl(url: str, timeout: int = 30, insecure: bool = False) -> tuple[int, str, str]:
    """Return (http_status, body, error). http_status 0 means transport failure.

    `insecure` skips TLS verification. Set it per-institution in surse.json, never
    globally — a broken cert on one venue is not a reason to stop checking the rest.
    """
    cmd = ["curl", "-sS", "-L", "--compressed", "-m", str(timeout),
           "-A", UA, "-w", "\n__HTTP_STATUS__%{http_code}"]
    if insecure:
        cmd.append("-k")
    proc = subprocess.run(cmd + [url], capture_output=True, text=True)
    if proc.returncode != 0:
        why = CURL_ERRORS.get(proc.returncode, "see `man curl` EXIT CODES")
        err = proc.stderr.strip().splitlines()
        return 0, "", f"curl exit {proc.returncode}: {why} | {err[0] if err else ''}"
    body, _, status = proc.stdout.rpartition("\n__HTTP_STATUS__")
    return int(status or 0), body, ""


def _goto_render(page, url: str) -> tuple[str, str]:
    """Load `url` on a live Playwright page and return (rendered_dom, error)."""
    try:
        page.goto(url, wait_until="networkidle", timeout=60_000)
        page.wait_for_timeout(1500)  # late calendar widgets
        return page.content(), ""
    except Exception as exc:
        return "", f"playwright: {str(exc)[:200]}"


def fetch_rendered(url: str, shot_path: pathlib.Path | None = None) -> tuple[str, str]:
    """Return the RENDERED DOM after JS execution, via Playwright.

    Standalone variant: launches its own Chromium. Inside teatre.py's harvest,
    browser pages go through _harvest_browser_job instead, which shares ONE
    browser across all of them — launch+teardown costs seconds each.

    This is the correct escalation for a JS-driven page — not a screenshot.
    page.content() gives fully hydrated markup: real text, real hrefs, real
    structure. A screenshot gives pixels, from which you would then try to
    recover the structure you already had. The full-page PNG is captured as
    human-auditable evidence, NOT as the extraction input.

        pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return "", "playwright not installed: pip install playwright && playwright install chromium"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(user_agent=UA,
                                    viewport={"width": 1440, "height": 2400})
            html, err = _goto_render(page, url)
            if not err and shot_path:
                page.screenshot(path=str(shot_path), full_page=True)
            browser.close()
        return html, err
    except Exception as exc:
        return "", f"playwright: {str(exc)[:200]}"


def month_mentioned(html: str, year: int, month: int) -> bool:
    low = html.lower()
    return (any(m in low for m in RO_MONTHS[month])
            or f"{year}-{month:02d}-" in low)


# ================================================================= stage 0
def preflight(verbose: bool = True, concurrency: int = 4) -> tuple[int, dict[str, str]]:
    """Prove the environment works, per institution.

    A single unreachable venue is DEGRADED, not fatal — the run continues without
    it and says so in the output. Only a systemic failure (missing tools, or
    nothing at all reachable) aborts. Aborting the whole job because one of twelve
    sources is down is the exact all-or-nothing trap this project exists to avoid.

    The probes (DNS + HTTP, one per institution) run in a thread pool; the
    report keeps source order so the output stays diffable run to run.
    """
    t0 = time.monotonic()
    if verbose:
        print("STAGE 0 — preflight\n")
    states: dict[str, str] = {}

    tools_ok = True
    for tool in ("curl", "python3"):
        found = subprocess.run(["which", tool], capture_output=True).returncode == 0
        tools_ok &= found
        if verbose:
            print(f"  {'OK ' if found else 'MISSING'}  {tool}")
    if verbose:
        print()

    today = dt.date.today()

    def check(inst: dict):
        """One institution's DNS + HTTP probe. Runs in the thread pool."""
        slug = inst["slug"]
        url = fmt_url(inst["calendar_url"], today.year, today.month)
        host = host_of(url)
        try:
            dns = f"DNS {socket.gethostbyname(host)}"
        except OSError as exc:
            return (slug, host, None, f"DNS FAIL ({exc})", "DNS_FAIL", 0, "")
        status, _, err = curl(url, 20, insecure=inst.get("insecure", False))
        if status in (200, 301, 302):
            state = "OK"
        elif "exit 60" in err or "exit 35" in err:
            state = "TLS_FAIL"
        elif status == 0:
            state = "UNREACHABLE"
        else:
            state = f"HTTP_{status}"
        return (slug, host, dns, "", state, status, err)

    insts = load_sources()
    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
        checks = list(ex.map(check, insts))

    for slug, host, dns, dns_err, state, status, err in checks:
        states[slug] = state
        if verbose:
            if dns is None:
                print(f"  FAIL  {slug:<12} {host:<30} {dns_err}")
            else:
                mark = {"OK": "OK  "}.get(state, "WARN")
                print(f"  {mark}  {slug:<12} {host:<30} {dns:<20} "
                      f"HTTP {status or '-'} {err}")

    reachable = sum(1 for s in states.values() if s == "OK")
    degraded = [k for k, v in states.items() if v != "OK"]

    if verbose:
        print(f"\npreflight: {reachable}/{len(states)} reachable")
        if degraded:
            print("  degraded (run continues, marked in output): "
                  + ", ".join(f"{k}={states[k]}" for k in degraded))

    off = [i["slug"] for i in load_sources(include_disabled=True)
           if not i.get("enabled", True)]
    if verbose:
        stage_done("preflight", t0,
                   f"{reachable}/{len(states)} sources reachable"
                   + (f", degraded: {','.join(degraded)}" if degraded else "")
                   + (f", excluded: {','.join(off)}" if off else ""))
    if not tools_ok:
        if verbose:
            print("  ABORT: required tools missing.")
        return 1, states
    if reachable == 0:
        if verbose:
            print("  ABORT: nothing reachable — this is your network, not the sites.\n"
                  "         Check connectivity before diagnosing anything else.")
        return 1, states
    return 0, states


# ================================================================= stage 2
SECTION_HINT = re.compile(r"program|calendar|spectacol|repertoriu|stagiun|eveniment|afis", re.I)


def discover(period: Period) -> int:
    """Stage 1b — mine RAW snapshot HTML for candidate programme URLs.

    This exists because of a real failure: bulandra's calendar_url was set to the
    homepage (stale April listings, 0 rows extracted) while /program/ was linked
    from that very homepage. The original discovery pass used a markdown renderer
    that dropped <a href> elements, and its silence was misread as "no such page".
    A lossy view returning nothing is not evidence that nothing is there.
    Raw HTML is the only honest source for link discovery.
    """
    t0 = time.monotonic()
    print(f"STAGE 1b — discover programme URLs from snapshots [{period.label}]\n")
    day = SNAPDIR / period.label
    if not day.exists():
        sys.exit(f"no snapshots at {day} — run harvest first")

    configured = {i["slug"]: [i["calendar_url"]] + i.get("alt_urls", [])
                  for i in load_sources()}
    found_total = mismatch = 0

    for hp in sorted(day.glob("*.html")):
        slug = hp.stem.partition("__")[0]
        raw = hp.read_text(encoding="utf-8", errors="replace")
        if not raw:
            continue
        cm = CANONICAL_RE.search(raw)
        base = cm.group(1) if cm else ""

        def norm(u: str) -> str:
            """Absolutise and strip query/trailing slash, so /program/ and
            https://host/program/ compare equal instead of looking like news."""
            u = u.strip()
            if u.startswith("//"):
                u = "https:" + u
            elif u.startswith("/"):
                u = base + u
            return SCHEME_RE.sub("", QUERY_HASH_RE.sub("", u)).rstrip("/").lower()

        own = SCHEME_RE.sub("", base).lower()
        cands, seen = [], set()
        for href in HREF_RE.findall(raw):
            if not SECTION_HINT.search(href) or ASSET_EXT_RE.search(href):
                continue
            # Third-party ticket sites are not official sources — surse.json's
            # policy forbids them, so they are noise here, not discoveries.
            if href.startswith("http") and own and not norm(href).startswith(own.split("/")[0]):
                continue
            path = SCHEME_HOST_RE.sub("", href)
            # Depth 4, not 2. Masca's WordPress lives under /teatru/, so every real
            # page is depth 3 — a depth>2 filter silently dropped ALL of them and
            # printed "no candidate links — likely JS-rendered", which was false.
            # Tune filters to the messiest site, not the tidiest.
            if WP_NOISE_RE.search(href):
                continue
            key = norm(href)
            if path.count("/") > 4 or len(path) > 60 or key in seen or not key:
                continue
            seen.add(key)
            cands.append(href)
        found_total += len(cands)

        known = {norm(u) for u in configured.get(slug, [])}
        news = [c for c in cands if norm(c) not in known]
        flag = ""
        if news:
            mismatch += 1
            flag = "  <-- NOT in surse.json"
        print(f"  {slug:<12}{flag}")
        for c in cands[:6]:
            star = " *" if c in news else "  "
            print(f"     {star} {c}")
        if not cands:
            print("        (no candidate links — likely JS-rendered)")

    stage_done("discover", t0,
               f"{found_total} candidate URL(s), {mismatch} site(s) with links "
               f"missing from surse.json")
    if mismatch:
        print("     Lines marked * are reachable programme pages you are NOT fetching.\n"
              "     Check them by hand, then update surse.json.")
    return 0


def parts_for(inst: dict, period: Period) -> list[tuple[str, list[str]]]:
    """(part_id, candidate_urls) for one institution over the period.

    Month-parameterised calendars get one part per month in the range. Everything
    else gets a single 'all' part — those pages show a rolling window that we
    filter by date downstream.
    """
    urls = [inst["calendar_url"]] + inst.get("alt_urls", [])
    if not is_month_parameterised(inst["calendar_url"]):
        return [("all", [fmt_url(u, period.start.year, period.start.month) for u in urls])]
    return [(f"{y}-{m:02d}", [fmt_url(u, y, m) for u in urls])
            for y, m in period.months()]


def harvest(period: Period, concurrency: int = 4) -> int:
    """Fetch each calendar and write raw HTML to disk. The snapshot is evidence.

    Curl parts run in a thread pool (curl is a subprocess, so threads really do
    overlap); browser parts share ONE Chromium — launching per page cost seconds
    each, and Playwright's sync API must stay on a single thread anyway.
    """
    t0 = time.monotonic()
    print(f"STAGE 2 — harvest for {period.human()}  [{period.label}]\n")
    day = SNAPDIR / period.label
    day.mkdir(parents=True, exist_ok=True)

    curl_jobs, browser_jobs = [], []
    for inst in load_sources():
        for part, candidates in parts_for(inst, period):
            (browser_jobs if inst.get("strategy") == "browser"
             else curl_jobs).append((inst, part, candidates))

    records: dict[tuple[str, str], dict] = {}
    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
        futs = [ex.submit(_harvest_curl_job, day, inst, part, cands, period)
                for inst, part, cands in curl_jobs]
        if browser_jobs:
            futs.append(ex.submit(_harvest_browser_job, day, browser_jobs, period))
        for fut in as_completed(futs):
            for rec in fut.result():
                records[(rec["slug"], rec["part"])] = rec

    n_ok = n_fail = raw_bytes = cond_bytes = n_canary = 0
    for inst in load_sources():
        insecure = inst.get("insecure", False)
        for part, _ in parts_for(inst, period):
            rec = records[(inst["slug"], part)]
            body, cond = rec["body"], rec["cond"]
            if rec["status"] == 200 and body:
                n_ok += 1
            else:
                n_fail += 1
            raw_bytes += len(body); cond_bytes += len(cond)
            hint = (f"-> {len(cond):>6}B condensed "
                    f"| dates raw={rec['n_raw']} kept={rec['n_cond']}" if body else "")
            flag = " [TLS UNVERIFIED]" if insecure else ""
            print(f"  {rec['stem']:<24} HTTP {rec['status'] or '-':<4} {len(body):>7}B"
                  f"{flag} {rec['err']} {hint}")
            if rec["lost"]:
                n_canary += 1
                print(f"      !! CANARY: {len(rec['lost'])} in-period date(s) present in the "
                      f"fetched HTML but MISSING from the payload sent to the model.")
                print(f"         lost: {', '.join(rec['lost'][:10])}"
                      + (f" (+{len(rec['lost']) - 10} more)" if len(rec['lost']) > 10 else ""))
                print(f"         This is a CONDENSER bug, not an empty venue. "
                      f"Inspect {rec['stem']}.html before believing any 'no schedule' result.")

    print(f"\nsnapshots -> {day}")
    stage_done("harvest", t0,
               f"{n_ok} fetched, {n_fail} failed · {raw_bytes/1e6:.2f} MB raw "
               f"-> {cond_bytes/1e3:.1f} KB condensed "
               f"({raw_bytes/max(cond_bytes,1):.0f}x, ~{cond_bytes/3.5:,.0f} tok)"
               + (f" · !! {n_canary} CANARY FAILURE(S) — dates lost in condensing"
                  if n_canary else " · canary clean"))
    return 0


def _harvest_curl_job(day: pathlib.Path, inst: dict, part: str,
                      candidates: list[str], period: Period) -> list[dict]:
    """Worker: try each candidate URL until one serves the page, then write the
    snapshot artifacts. Files are per-stem, so workers never touch one path."""
    status, body, err, url = 0, "", "no url tried", ""
    for url in candidates:
        status, body, err = curl(url, insecure=inst.get("insecure", False))
        if status == 200 and body:
            break  # first working source wins; alts are fallbacks
    return [_finalize_part(day, inst, part, url, status, body, err, period)]


def _harvest_browser_job(day: pathlib.Path, browser_jobs: list[tuple],
                         period: Period) -> list[dict]:
    """Worker: every browser-strategy part under ONE Chromium instance.

    Playwright's sync API is not safe across threads, so the whole batch lives
    on this single worker — running alongside the curl workers in the pool.
    """
    out = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        for inst, part, cands in browser_jobs:
            out.append(_finalize_part(day, inst, part, cands[0], 0, "",
                                      "playwright not installed: pip install playwright "
                                      "&& playwright install chromium", period))
        return out
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(user_agent=UA, viewport={"width": 1440, "height": 2400})
        for inst, part, cands in browser_jobs:
            status, body, err, url = 0, "", "no url tried", ""
            for url in cands:
                body, err = _goto_render(page, url)
                status = 200 if body else 0
                if status == 200 and body:
                    break
            if body:
                try:
                    page.screenshot(path=str(day / f"{inst['slug']}__{part}.png"),
                                    full_page=True)
                except Exception:
                    pass  # the evidence photo must not sink a good DOM
            out.append(_finalize_part(day, inst, part, url, status, body, err, period))
        browser.close()
    return out


def _finalize_part(day: pathlib.Path, inst: dict, part: str, url: str,
                   status: int, body: str, err: str, period: Period) -> dict:
    """Write .html/.txt/.meta.json for one part; return what the report needs."""
    stem = f"{inst['slug']}__{part}"
    (day / f"{stem}.html").write_text(body, encoding="utf-8")
    # The condensed text is the stage-3 payload AND the stage-4 evidence.
    # Deterministically derived from the snapshot, and small enough to read.
    cond = condense(body, period) if body else ""
    n_raw, n_cond, lost = canary(body, cond, period) if body else (0, 0, [])
    (day / f"{stem}.txt").write_text(cond, encoding="utf-8")
    (day / f"{stem}.meta.json").write_text(json.dumps({
        "slug": inst["slug"], "name": inst["name"], "part": part,
        "url": url, "http_status": status, "error": err, "bytes": len(body),
        "tls_verified": not inst.get("insecure", False),
        "condensed_bytes": len(cond),
        "period": period.label,
        "period_start": period.start.isoformat(),
        "period_end": period.end.isoformat(),
        "fetched_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "strategy": inst["strategy"],
        "status": "FETCHED" if status == 200 and body else
                  ("NEEDS_BROWSER" if inst["strategy"] == "browser"
                   else "FETCH_FAILED"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"slug": inst["slug"], "part": part, "stem": stem, "status": status,
            "body": body, "err": err, "cond": cond,
            "n_raw": n_raw, "n_cond": n_cond, "lost": lost}


# ================================================================= structure
def extract_structured(html: str) -> list[tuple[str, str]]:
    """Pull machine-readable payloads out of <script> tags.

    This is where schedules often live, and stripping it was a bug:
      - <script type="application/ld+json">  schema.org Event
      - <script id="__NEXT_DATA__">          Next.js page props
      - <script type="application/json">     generic embedded state
    """
    found = []
    for m in SCRIPT_TAG_RE.finditer(html):
        attrs, body = m.group(1).lower(), m.group(2).strip()
        if not body:
            continue
        if "ld+json" in attrs:
            kind = "ld+json"
        elif "__next_data__" in attrs:
            kind = "__NEXT_DATA__"
        elif "application/json" in attrs:
            kind = "embedded-json"
        else:
            continue
        found.append((kind, body))
    return found


def count_schema_events(blocks: list[tuple[str, str]]) -> int:
    """Count schema.org Event nodes ONLY.

    Nearly every WordPress site emits ld+json via Yoast SEO — WebPage,
    BreadcrumbList, WebSite, Organization. That is site chrome, not a schedule.
    Counting raw ld+json blocks made 9 of 12 venues look RICH when only one
    actually publishes structured events.
    """
    n = 0
    for kind, body in blocks:
        if kind != "ld+json":
            continue
        try:
            data = json.loads(body)
        except json.JSONDecodeError:
            continue
        graph = data.get("@graph", data) if isinstance(data, dict) else data
        for node in (graph if isinstance(graph, list) else [graph]):
            if isinstance(node, dict):
                t = node.get("@type")
                t = t[0] if isinstance(t, list) and t else t
                if isinstance(t, str) and "Event" in t:
                    n += 1
    return n


def summarise_structured(blocks: list[tuple[str, str]], period: Period) -> str:
    """Keep structured blocks that plausibly concern the period."""
    toks = []
    for y, m in period.months():
        toks += RO_MONTHS[m] + [f"-{m:02d}-", f"{y}-{m}-"]
    keep = []
    for kind, body in blocks:
        low = body.lower()
        if any(t in low for t in toks) or '"@type":"event"' in low.replace(" ", ""):
            keep.append(f"<!-- STRUCTURED DATA [{kind}] -->\n{body[:60_000]}")
    return "\n\n".join(keep)


def period_signals(text: str, period: Period) -> set[str]:
    """Normalised in-period date tokens found in `text`.

    Deliberately narrow: only dates that fall INSIDE the requested period count.
    A generic date regex matches SVG path coordinates, phone numbers and "WCAG
    2.1"; those must not register as schedule signal or the canary below cries
    wolf on every page.
    """
    found: set[str] = set()
    year_of_month = {m: y for y, m in period.months()}

    def add(y: int | str, m: int | str, d: int | str) -> None:
        """Canonicalise to a real date, then keep it only if it is in period.

        Every format must normalise to the SAME key. An earlier version emitted
        '2026-09-12' for ISO and '12.09' for dotted, so the identical date read as
        two different signals and the comparison was meaningless.
        """
        try:
            when = dt.date(int(y), int(m), int(d))
        except ValueError:
            return
        if period.contains(when):
            found.add(when.isoformat())

    # (?!\d), NOT \b. In "2026-09-05T19:00" the character after "05" is "T" — a
    # word character — so a trailing \b does not match and 22 ISO stamps read as 2.
    # This is the same boundary bug that made me declare Bulandra empty; the guard
    # against it shipped containing it.
    for y, m, d in ISO_DATE_RE.findall(text):
        add(y, m, d)

    for d, m in DOT_DATE_RE.findall(text):
        if int(m) in year_of_month:
            add(year_of_month[int(m)], m, d)

    for y, m in period.months():
        for name in RO_MONTHS[m]:
            for d in MONTH_NAME_DATE_RE[name].findall(text):
                add(y, m, d)
    return found


def canary(raw: str, condensed: str, period: Period) -> tuple[int, int, list[str]]:
    """Did condensing LOSE in-period dates that were present in the raw HTML?

    This is the guard that five separate bugs needed and did not have. Every one
    of them had the same signature — data present in the fetched bytes, absent
    from what reached the model, reported upward as "this venue has no schedule":

      1. bulandra   calendar_url pointed at a stale homepage
      2. metropolis 59 links destroyed by a <url> marker the tag-stripper ate
      3. masca      every URL dropped by a depth>2 filter
      4. nottara    dates split across source lines, then buried in boilerplate
      5. bulandra   schedule inside a plain <script>, stripped wholesale

    An empty result is a claim about the world. This checks the claim.
    """
    # <svg> path data is a minefield of decimal pairs — "6.73-18.09,14.42" reads as
    # 18 September to any date regex. Vector coordinates are never schedule data,
    # so exclude them from the raw side. <script> is NOT excluded: Bulandra's whole
    # programme lives in one, and catching that is the entire point.
    raw_content = SVG_STYLE_RE.sub(" ", raw)
    in_raw = period_signals(raw_content, period)
    in_cond = period_signals(condensed, period)
    missing = in_raw - in_cond

    # Threshold, so a single stray numeric coincidence doesn't cry wolf every run.
    # Fires when the payload lost EVERYTHING, or lost a substantial share.
    significant = bool(missing) and (
        not in_cond or len(missing) >= max(3, 0.3 * len(in_raw)))
    return len(in_raw), len(in_cond), sorted(missing) if significant else []


def salvage_script_schedules(raw: str, period: Period, radius: int = 420) -> str:
    """Recover schedules embedded in PLAIN inline <script> payloads.

    Bulandra publishes its entire programme as a JS object inside
    <script id="wcs-main-js-extra">:

        "start":"2026-09-05T19:00:00+00:00","end":"2026-09-05T20:50:00+00:00",
        "permalink":"https://www.bulandra.ro/program/family-exe-1-2-2-2/"

    That script has no `type` attribute, so it is neither ld+json nor
    __NEXT_DATA__ nor application/json — every typed-payload salvage missed it,
    and the generic <script> strip then deleted it. Result: a page carrying 44
    dated performances extracted as zero.

    Rather than parse each plugin's schema, cut a window around every ISO
    datetime that falls inside the period and hand the model the neighbourhood.
    Format-agnostic, so it works for any calendar plugin that emits ISO stamps.
    """
    wanted = {f"{y}-{m:02d}" for y, m in period.months()}
    out, seen = [], set()
    for m in SCRIPT_BODY_RE.finditer(raw):
        body = m.group(1)
        if not any(w in body for w in wanted):
            continue
        # Time is OPTIONAL. Grivița53 emits bare '2026-09-07' with no T-component;
        # requiring one skipped its whole schedule.
        for d in STAMP_RE.finditer(body):
            if d.group(0)[:7] not in wanted:
                continue
            lo = max(0, d.start() - radius)
            if any(abs(lo - s) < radius // 2 for s in seen):
                continue                      # overlapping window already taken
            seen.add(lo)
            win = body[lo:d.end() + radius]
            # Unescape the JS/JSON encoding so titles are readable text.
            win = UNICODE_ESC_RE.sub(lambda x: chr(int(x.group(1), 16)), win)
            win = win.replace("\\/", "/").replace('\\"', '"')
            win = TAG_STRIP_RE.sub(" ", win)
            out.append(WS_RE.sub(" ", win).strip())
    if not out:
        return ""
    return "<!-- SCHEDULE FOUND IN INLINE SCRIPT -->\n" + "\n---\n".join(out[:60])


# Legal/consent/a11y furniture that every Romanian theatre site carries. It is
# never schedule content, and it is what filled the payload when date matching
# failed — 1.5 KB of Bulandra "condensed" was 100% accessibility statement.
BOILERPLATE = re.compile(
    r"WCAG|W3C|accessib|screen-reader|GDPR|cookie|consimțăm|confidenţial|"
    r"confidențial|Directiva|Regulament|politica de|newsletter|"
    r"Compliance status|disabilit", re.I)


def condense(raw: str, period: Period, ctx: int = 3) -> str:
    """Reduce a snapshot to schedule-bearing lines. THE payload for stage 3.

    Measured over the 11 real snapshots: 1.79 MB raw -> 732 KB via trim_html ->
    18 KB condensed. That is ~209k tokens down to ~5k, a 40x cut, and it is the
    single biggest lever on extraction speed — far bigger than which model runs.
    Sending 60k tokens of <div class="col-12 col-md-9"> to any model is the bug;
    swapping models to make the bug faster is not the fix.

    Links are preserved inline as `text <url>` so the model can still emit `link`.
    Only lines within `ctx` of a date match are kept.
    """
    h = NOISE_BLOCK_RE.sub(" ", raw)
    h = COMMENT_RE.sub(" ", h)
    # Preserve links as `text [url]`. NOT `<url>` — angle brackets get eaten by the
    # generic <[^>]+> strip three lines below, which silently destroyed all 59
    # anchors on the Metropolis page and left the model inventing every link value.
    # Surface machine-readable date/time attributes as visible text BEFORE tags are
    # stripped. Nottara carries data-fulldate="2026-09-12T19:30:00+00:00" and
    # data-venue on every row — an exact, unambiguous schedule that attribute
    # stripping was throwing away in favour of guessing at prose.
    def _surface(m: re.Match) -> str:
        tag = m.group(0)
        vals = DATA_ATTR_VAL_RE.findall(tag)
        # AFTER the tag's closing '>', not inside it — text injected within the tag
        # is removed along with the tag by the generic stripper further down.
        return tag + " " + " ".join(vals) + " " if vals else tag

    h = DATA_ATTR_TAG_RE.sub(_surface, h)
    h = ANCHOR_RE.sub(r"\2 [\1]", h)

    # Kill SOURCE formatting first. Nottara writes each part of a date on its own
    # source line (<span>Sâm</span>\n<span>12</span>\n<span>Sep</span>), so any
    # line-based date match failed on "12" and "Sep" separately. Collapse all
    # whitespace, THEN insert breaks at block boundaries only — inline elements
    # must not fragment a date.
    h = NBSP_RE.sub(" ", h)
    h = WS_RE.sub(" ", h)
    h = BLOCK_BREAK_RE.sub("\n", h)
    h = TD_BREAK_RE.sub(" | ", h)
    h = TAG_STRIP_RE.sub(" ", h)
    h = "\n".join(WS_LINE_RE.sub(" ", ln).strip() for ln in h.split("\n"))
    h = re.sub(r"\n{2,}", "\n", h).strip()

    lines = [ln for ln in h.split("\n") if not BOILERPLATE.search(ln)]
    keep: set[int] = set()
    for i, ln in enumerate(lines):
        if DATE_PAT.search(ln):
            keep.update(range(max(0, i - ctx), min(len(lines), i + ctx + 1)))

    body = "\n".join(lines[i] for i in sorted(keep)) if keep else ""

    # Safety net: never ship a payload with no sign of the target period. If the
    # date filter produced nothing relevant, it is the FILTER that failed, not the
    # site — fall back to the full text rather than confidently sending cookie
    # notices and letting the model conclude NO_SCHEDULE_FOUND.
    months = [m for _, mo in period.months() for m in RO_MONTHS[mo]]
    if not any(m in body.lower() for m in months) and not re.search(r"\b\d{1,2}[:.]\d{2}\b", body):
        body = "\n".join(lines)[:40_000]

    head = "\n\n".join(x for x in (summarise_structured(extract_structured(raw), period),
                                   salvage_script_schedules(raw, period)) if x)
    return f"{head}\n\n<!-- TEXT -->\n{body}" if head else body


DATE_PAT = re.compile(
    r"\b\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?\b"
    r"|\b\d{1,2}\s*(?:" + "|".join(sum(RO_MONTHS.values(), [])) + r")\b"
    r"|\b\d{4}-\d{1,2}-\d{1,2}\b", re.I)


# ================================================================= stage 2b
def diagnose(period: Period) -> int:
    """Report what each snapshot ACTUALLY contains, before redesigning around it."""
    t0 = time.monotonic()
    print(f"STAGE 2b — diagnose snapshots for {period.human()}\n")
    day = SNAPDIR / period.label
    if not day.exists():
        sys.exit(f"no snapshots at {day} — run harvest first")

    print(f"  {'snapshot':<24} {'raw':>8} {'text':>7} {'ld':>3} {'nxt':>4} "
          f"{'dates':>6} {'luna':>5}  verdict")
    print("  " + "-" * 88)

    advice, seen_months = [], []
    for hp in sorted(day.glob("*.html")):
        stem = hp.stem
        slug, _, part = stem.partition("__")
        raw = hp.read_text(encoding="utf-8", errors="replace")
        blocks = extract_structured(raw)
        ld = count_schema_events(blocks)
        nxt = any(k == "__NEXT_DATA__" for k, _ in blocks)
        text = TAG_STRIP_RE.sub(" ", SCRIPT_STYLE_SVG_RE.sub(" ", raw))
        text = WS_RE.sub(" ", text)
        ndates = len(set(DATE_PAT.findall(text)))
        y, m = ((period.start.year, period.start.month) if part == "all"
                else (int(part[:4]), int(part[5:])))
        in_month = month_mentioned(raw, y, m)
        seen_months.append(in_month)

        if ndates >= 5 or ld or nxt:
            verdict = "RICH — parse the HTML/JSON"
        elif ndates == 0 and len(text) < 2_000:
            verdict = "EMPTY — likely JS-rendered; try browser"
            advice.append((slug, "browser"))
        elif ndates == 0:
            verdict = "NO DATES — wrong URL, or dates are images"
            advice.append((slug, "check-url"))
        else:
            verdict = f"THIN — {ndates} date(s); check coverage"
            advice.append((slug, "check-url"))

        print(f"  {stem:<24} {len(raw):>8} {len(text):>7} {ld:>3} "
              f"{'yes' if nxt else '  -':>4} {ndates:>6} "
              f"{'yes' if in_month else '  -':>5}  {verdict}")

    print("\n  ld  = schema.org *Event* nodes (exact dates; SEO chrome not counted)")
    print("  nxt = Next.js __NEXT_DATA__ payload present")
    print("  luna= that part's month name appears anywhere in the raw HTML")

    quiet = sum(1 for x in seen_months if not x)
    if seen_months and quiet >= len(seen_months) / 2:
        names = "/".join(RO_MONTHS[m][0] for _, m in period.months())
        print(f"\n  NOTE: {quiet}/{len(seen_months)} snapshots never mention {names}. "
              "Either the period is out\n        of season (Romanian theatres largely go "
              "dark mid-July to early September)\n        or it is beyond the published "
              "horizon. Thin results here are the venues' doing.")

    rich = len(seen_months) - len(set(a[0] for a in advice))
    stage_done("diagnose", t0,
               f"{len(seen_months)} snapshot(s): {rich} parseable, "
               f"{len(set(a[0] for a in advice))} need attention")
    if advice:
        print("\n  follow-ups:")
        for slug, what in dict.fromkeys(advice):
            print(f"    {slug:<12} "
                  + ("set strategy=browser in surse.json" if what == "browser"
                     else "calendar_url may be wrong — re-check by hand"))
    else:
        print("\n  All snapshots carry parseable data. No browser needed.")
    return 0


# ================================================================= stage 3
# ================================================= deterministic extractors
# Where a venue publishes machine-readable timestamps, running them through a
# model to turn "2026-09-12T19:30:00" into {"data": "...", "ora": "..."} is paying
# an LLM to do strptime. These extractors are free, instant, and cannot
# hallucinate. The model's real job is the messy prose the others can't handle.

def _iso(s: str) -> tuple[dt.date, str] | None:
    """Parse a tolerant ISO stamp. Odeon emits '2026-9-12T20:00+0:00' — month and
    offset unpadded, which is not valid ISO 8601 and breaks fromisoformat()."""
    m = re.match(r"\s*(\d{4})-(\d{1,2})-(\d{1,2})(?:[T ](\d{1,2}):(\d{2}))?", s)
    if not m:
        return None
    y, mo, d, hh, mm = m.groups()
    try:
        return dt.date(int(y), int(mo), int(d)), (f"{int(hh):02d}:{mm}" if hh else "")
    except ValueError:
        return None


def rows_from_jsonld(raw: str, period: Period) -> list[dict]:
    """schema.org Event nodes. Odeon publishes 13 of them."""
    out = []
    for kind, body in extract_structured(raw):
        if kind != "ld+json":
            continue
        try:
            data = json.loads(body)
        except json.JSONDecodeError:
            continue
        graph = data.get("@graph", data) if isinstance(data, dict) else data
        for node in (graph if isinstance(graph, list) else [graph]):
            if not isinstance(node, dict):
                continue
            t = node.get("@type")
            t = t[0] if isinstance(t, list) and t else t
            if not (isinstance(t, str) and "Event" in t) or not node.get("startDate"):
                continue
            parsed = _iso(str(node["startDate"]))
            if not parsed or not period.contains(parsed[0]):
                continue
            loc = node.get("location")
            out.append({
                "data": parsed[0].isoformat(), "ora": parsed[1] or "—",
                "spectacol": str(node.get("name", "—")).strip(),
                "sala": (loc.get("name") if isinstance(loc, dict) else None) or "—",
                "link": str(node.get("url") or ""),
                # Must be a literal substring of the condensed payload, because
                # verify checks it there — the raw stamp satisfies that.
                "citat_sursa": str(node["startDate"]),
            })
    return out


def rows_from_data_attrs(raw: str, period: Period) -> list[dict]:
    """data-fulldate / data-venue on the row container. Nottara's pattern."""
    out = []
    for m in DATA_FULLDATE_RE.finditer(raw):
        parsed = _iso(m.group(1))
        if not parsed or not period.contains(parsed[0]):
            continue
        window = raw[m.start():m.start() + 2500]
        venue = DATA_VENUE_RE.search(window)
        link = SPECTACOL_HREF_RE.search(window)
        title = SPECTACOL_TITLE_RE.search(window)
        time_txt = HOUR_TEXT_RE.search(window)
        out.append({
            "data": parsed[0].isoformat(),
            "ora": parsed[1] or (time_txt.group(1) if time_txt else "—"),
            "spectacol": (title.group(1).strip() if title else "—"),
            "sala": UPPER_SPACE_RE.sub(" ", venue.group(1)).strip() if venue else "—",
            "link": link.group(1) if link else "",
            "citat_sursa": m.group(1),
        })
    return out


def rows_from_inline_json(raw: str, period: Period) -> list[dict]:
    """`"start":"<ISO>"` inside a plain <script> payload. Bulandra's wcs plugin."""
    out = []
    for m in INLINE_START_RE.finditer(raw):
        parsed = _iso(m.group(1))
        if not parsed or not period.contains(parsed[0]):
            continue
        window = raw[m.start():m.start() + 2000]
        perma = PERMALINK_RE.search(window)
        subj = SUBJECT_RE.search(window)
        sala = SALA_RE.search(window)
        title = "—"
        if subj:
            title = urllib.parse.unquote_plus(subj.group(1)).strip()
        elif perma:
            slug = perma.group(1).rstrip("/").split("/")[-1].split("?")[0]
            title = SLUG_NUMS_RE.sub("", slug).replace("-", " ").upper()
        out.append({
            "data": parsed[0].isoformat(), "ora": parsed[1] or "—",
            "spectacol": title, "sala": sala.group(0) if sala else "—",
            "link": (perma.group(1).replace("\\/", "/") if perma else ""),
            "citat_sursa": m.group(1),
        })
    return out


def rows_from_fullcalendar(raw: str, period: Period) -> list[dict]:
    """FullCalendar month grid (TNB). Events are located by COLUMN, not by attribute.

    The library renders each week as a "content skeleton": a <thead> carrying one
    data-date per weekday column, then <tbody> rows whose <td>s line up under those
    columns. An event's own markup contains no date at all — its position does.
    Empty leading cells are collapsed into <td rowspan="5"></td> spacers, so column
    index must be tracked with rowspan occupancy or every event lands on the wrong
    day.

    Note also that data-date appears on EVERY cell including empty ones — 42 per
    month. Reading those directly would manufacture a phantom show for every square
    on the calendar.
    """
    rows: list[dict] = []
    # Segment by marker, not by a non-greedy </div> — the skeleton contains dozens
    # of nested divs, so `(.*?)</div>` captures only the first few bytes of it.
    for chunk in raw.split('<div class="fc-content-skeleton">')[1:]:
        cuts = [x for x in (chunk.find('class="fc-bg"'),
                            chunk.find("fc-row fc-week")) if x > 0]
        skel = chunk[:min(cuts)] if cuts else chunk
        head, _, body = skel.partition("</thead>")
        dates = DATA_DATE_RE.findall(head)
        if not dates:
            continue
        occupied: dict[int, int] = {}
        for tr in TR_RE.findall(body):
            col = 0
            for cell in TD_SPLIT_RE.split(tr):
                if not cell.lstrip().startswith("<td"):
                    continue
                attrs = cell[:cell.find(">") + 1]
                while occupied.get(col, 0) > 0:
                    occupied[col] -= 1
                    col += 1
                cm = COLSPAN_RE.search(attrs)
                span = int(cm.group(1)) if cm else 1
                rm = ROWSPAN_RE.search(attrs)
                rspan = int(rm.group(1)) if rm else 1

                if "fc-event-container" in attrs and col < len(dates):
                    when = dt.date.fromisoformat(dates[col])
                    if period.contains(when):
                        for ev in FC_EVENT_RE.findall(cell) or [cell]:
                            title = FC_TITLE_RE.search(ev)
                            hour = FC_HOUR_RE.search(ev)
                            loc = FC_LOC_RE.search(ev)
                            link = FC_LINK_RE.search(ev)
                            if not title:
                                continue
                            rows.append({
                                "data": when.isoformat(),
                                "ora": hour.group(1) if hour else "—",
                                "spectacol": TAG_STRIP_RE.sub("", title.group(1)).strip(),
                                "sala": TAG_STRIP_RE.sub("", loc.group(1)).strip() if loc else "—",
                                "link": link.group(1) if link else "",
                                # The day cell's data-date IS in the payload, so verify
                                # can confirm this row against the snapshot.
                                "citat_sursa": dates[col],
                            })
                if rspan > 1:
                    for c in range(col, col + span):
                        occupied[c] = rspan - 1
                col += span
    return rows


DETERMINISTIC = {"jsonld": rows_from_jsonld, "data-attrs": rows_from_data_attrs,
                 "inline-json": rows_from_inline_json,
                 "fullcalendar": rows_from_fullcalendar}


def _raw_signals(raw: str) -> set[str]:
    """Cheap substring gates for the deterministic extractors.

    One lowercase pass; each extractor's regex is guaranteed to need its gate
    token, so a failed gate means an empty result — no full-page scan needed.
    """
    low = raw.lower()
    sig = set()
    if "ld+json" in low:
        sig.add("jsonld")
    if "data-fulldate" in low:
        sig.add("data-attrs")
    if '"start"' in low:
        sig.add("inline-json")
    if "fc-content-skeleton" in low:
        sig.add("fullcalendar")
    return sig


def deterministic_rows(raw: str, period: Period,
                       strategy: str = "auto") -> tuple[str, list[dict]]:
    """Try the configured strategy, or all of them. First non-empty wins.

    Returns (strategy_used, rows). ('', []) means fall through to the model.
    In auto mode each extractor is skipped when its signal gate fails, so a
    page that matches none is scanned once (lowercased) instead of four times.
    """
    names = list(DETERMINISTIC) if strategy in ("auto", "", None) else [strategy]
    gates: set[str] | None = None
    for name in names:
        fn = DETERMINISTIC.get(name)
        if not fn:
            continue
        if strategy in ("auto", "", None):
            if gates is None:
                gates = _raw_signals(raw)
            if name not in gates:
                continue
        try:
            rows = fn(raw, period)
        except Exception:
            rows = []
        if rows:
            # Same performance can appear twice in a payload; collapse.
            seen, uniq = set(), []
            for r in rows:
                key = (r["data"], r["ora"], r["spectacol"])
                if key not in seen:
                    seen.add(key)
                    uniq.append(r)
            return name, uniq
    return "", []


ROW_SCHEMA = {
    "type": "object",
    "properties": {
        "status": {"type": "string",
                   "enum": ["OK", "NOT_PUBLISHED", "NO_SCHEDULE_FOUND"]},
        "published_through": {"type": "string"},
        "rows": {"type": "array", "items": {
            "type": "object",
            "properties": {
                "data": {"type": "string"}, "ora": {"type": "string"},
                "spectacol": {"type": "string"}, "sala": {"type": "string"},
                "link": {"type": "string"}, "citat_sursa": {"type": "string"},
            },
            "required": ["data", "ora", "spectacol", "sala", "link", "citat_sursa"],
            "additionalProperties": False,
        }},
    },
    "required": ["status", "published_through", "rows"],
    "additionalProperties": False,
}

EXTRACT_PROMPT = """\
Below is the programme text of the official page for {name}, fetched {ts} from {url}.
Links appear inline as `title [https://...]`. Copy the `link` value verbatim from
the bracketed URL next to that performance. If there is none, use "".

Target period: {start} to {end} inclusive.

Return JSON only.

If the page lists performances anywhere in that period:
{{"status":"OK","rows":[{{"data":"YYYY-MM-DD","ora":"HH:MM","spectacol":"...",
"sala":"...","link":"...","citat_sursa":"<the date/time string exactly as it appears in the HTML>"}}]}}

Include EVERY performance whose date falls in the period, including multiple
shows on the same day. One row per performance, not per production.

If the page's coverage ends before the period starts:
{{"status":"NOT_PUBLISHED","published_through":"YYYY-MM-DD"}}

If the page has no parseable schedule:
{{"status":"NO_SCHEDULE_FOUND"}}

Set "published_through" to "" unless status is NOT_PUBLISHED.

Do not infer from other years. Use nothing outside this text.
`citat_sursa` MUST be a literal substring of the text below.

--- TEXT ---
{html}
"""


# urllib defaults to `User-Agent: Python-urllib/3.x`, which Cloudflare's browser
# integrity rule bans outright — the request is rejected with "error code: 1010"
# before it ever reaches the API, so a perfectly valid key looks like a 403 auth
# failure. Every real SDK identifies itself; so do we.
API_UA = "teatre.py/1.0 (+https://github.com/; python-urllib)"


def api_headers(extra: dict[str, str]) -> dict[str, str]:
    return {"Content-Type": "application/json", "Accept": "application/json",
            "User-Agent": API_UA, **extra}


class HttpStatusError(RuntimeError):
    """A non-2xx HTTP response, surfaced with its status code and short body.

    A subclass of RuntimeError so every existing `except Exception`/`except
    RuntimeError` handler keeps working; it just carries enough structure for
    callers to decide whether to retry (e.g. OpenRouter 429/5xx rate limits).
    """

    def __init__(self, code: int, reason: str, body: str,
                 retry_after: float | None = None) -> None:
        super().__init__(f"HTTP {code} {reason} — {body or '(empty body)'}")
        self.code = code
        self.reason = reason
        self.body = body
        self.retry_after = retry_after


def post_json(req: urllib.request.Request, timeout: int = 600) -> dict:
    """POST and decode, surfacing the API's OWN error message on failure.
    …
    urllib's HTTPError stringifies to a bare 'HTTP Error 403: Forbidden' while the
    body — which says *why* — sits unread on the exception. Discarding it cost a
    debugging round: eleven identical 403s with no indication whether the cause was
    the key, the model, or the tier. Always read the body.
    """
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as exc:
        try:
            body = exc.read().decode("utf-8", "replace").strip()[:400]
        except Exception:
            body = ""
        retry_after = None
        try:
            raw = exc.headers.get("Retry-After")
            if raw and raw.strip().replace(".", "", 1).lstrip("-+").isdigit():
                retry_after = float(raw.strip())
        except Exception:
            retry_after = None
        raise HttpStatusError(exc.code, exc.reason, body, retry_after) from None
    except urllib.error.URLError as exc:
        raise RuntimeError(f"network unreachable: {exc.reason}") from None


def call_ollama(prompt: str, model: str) -> str:
    host = os.environ.get("OLLAMA_HOST", "http://127.0.0.1:11434").rstrip("/")
    payload = json.dumps({
        "model": model, "prompt": prompt, "stream": False,
        "format": "json", "options": {"temperature": 0},
    }).encode()
    req = urllib.request.Request(f"{host}/api/generate", payload, api_headers({}))
    data = post_json(req)
    out = data.get("response")
    if not out:
        raise RuntimeError(f"empty response (done_reason={data.get('done_reason')})")
    return out


def call_cerebras(prompt: str, model: str) -> str:
    """Cerebras Inference — OpenAI-compatible, with strict JSON-schema enforcement.

    Free tier is 30K tokens/minute. That is why `condense` matters: the old
    trim_html payload for Odeon alone was ~58k tokens and would be rejected
    outright as a single request.
    """
    key = os.environ.get("CEREBRAS_API_KEY")
    if not key:
        raise RuntimeError("CEREBRAS_API_KEY not set")
    payload = json.dumps({
        "model": model, "temperature": 0,
        "max_completion_tokens": int(os.environ.get("TEATRE_MAX_TOKENS", 32768)),
        "messages": [{"role": "user", "content": prompt}],
        "response_format": {"type": "json_schema", "json_schema": {
            "name": "programme", "strict": True, "schema": ROW_SCHEMA}},
    }).encode()
    req = urllib.request.Request(
        "https://api.cerebras.ai/v1/chat/completions", payload,
        api_headers({"Authorization": f"Bearer {key.strip()}"}))
    data = post_json(req, timeout=300)

    choice = (data.get("choices") or [{}])[0]
    msg = choice.get("message") or {}
    content = msg.get("content")
    if content:
        return content

    # gpt-oss-120b is a REASONING model: it spends completion tokens on `reasoning`
    # first, and `content` appears only once that finishes. Cap it too low and you
    # get a message with `reasoning` and no `content` at all — which used to blow up
    # as a bare KeyError. Say what actually happened instead.
    fin = choice.get("finish_reason")
    if fin == "length":
        used = (data.get("usage") or {}).get("completion_tokens", "?")
        raise RuntimeError(
            f"hit max_completion_tokens ({used} used) before emitting content — "
            "reasoning consumed the budget; raise TEATRE_MAX_TOKENS or use a "
            "non-reasoning model")
    raise RuntimeError(f"no content in reply (finish_reason={fin}, "
                       f"message keys={sorted(msg)})")


def call_mimo(prompt: str, model: str) -> str:
    """MiMo-V2.5 (Xiaomi) via OpenRouter — OpenAI-compatible chat completions.

    Routed through OpenRouter's unified API so one key reaches many providers,
    and (the point the user asked for) MiMo's inference is slower than Cerebras',
    which makes the live run observable in the v2 visualizer instead of finishing
    in an instant.

    OpenRouter's free tier rate-limits hot models (HTTP 429 / 5xx). Those are
    transient, so a single call retries with backoff (honouring Retry-After when
    present) instead of silently dropping a venue, up to TEATRE_MAX_RETRIES.
    """
    key = os.environ.get("OPENROUTER_API_KEY")
    if not key:
        raise RuntimeError("OPENROUTER_API_KEY not set")
    payload = json.dumps({
        "model": model, "temperature": 0,
        "max_tokens": int(os.environ.get("TEATRE_MAX_TOKENS", 32768)),
        "messages": [{"role": "user", "content": prompt}],
        "response_format": {"type": "json_schema", "json_schema": {
            "name": "programme", "strict": True, "schema": ROW_SCHEMA}},
    }).encode()
    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions", payload,
        api_headers({"Authorization": f"Bearer {key.strip()}"}))

    attempts = max(1, int(os.environ.get("TEATRE_MAX_RETRIES", "4")))
    for attempt in range(1, attempts + 1):
        try:
            data = post_json(req, timeout=600)
            break
        except HttpStatusError as exc:
            retryable = exc.code in (429, 500, 502, 503, 504)
            if not retryable or attempt == attempts:
                raise
            wait = exc.retry_after if exc.retry_after else float(2 ** attempt)
            print(f"  [mimo] {model} transient HTTP {exc.code} — "
                  f"retry {attempt}/{attempts} in {wait:.0f}s", file=sys.stderr)
            time.sleep(min(wait, 60.0))

    choice = (data.get("choices") or [{}])[0]
    msg = choice.get("message") or {}
    content = msg.get("content")
    if content:
        return content
    fin = choice.get("finish_reason")
    if fin == "length":
        used = (data.get("usage") or {}).get("completion_tokens", "?")
        raise RuntimeError(
            f"hit max_completion_tokens ({used} used) before emitting content — "
            "raise TEATRE_MAX_TOKENS or use a non-reasoning model")
    raise RuntimeError(f"no content in reply (finish_reason={fin}, "
                       f"message keys={sorted(msg)})")


_JSON_DECODER = json.JSONDecoder()


def parse_json_loose(text: str) -> dict:
    """Models wrap JSON in prose or code fences. Recover the object, or fail loudly.

    Uses the C-implemented raw_decode instead of a character-by-character Python
    depth scan: same recovery semantics (first balanced object, skipping any
    stray braces before it), microseconds on a 50 KB reply.
    """
    text = FENCE_RE.sub("", text.strip())
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    start = text.find("{")
    while start >= 0:
        try:
            return _JSON_DECODER.raw_decode(text, start)[0]
        except json.JSONDecodeError:
            start = text.find("{", start + 1)
    raise ValueError(f"no JSON object in model output: {text[:200]!r}")


def extract(period: Period, backend: str = "none", model: str = "",
            concurrency: int = 2) -> int:
    """One model call per snapshot. Small context, one job, strict JSON out.

    Model calls are the slow stage (tens of seconds each) and independent per
    snapshot, so they run with bounded concurrency. The default is 2, not more:
    API tiers limit tokens/minute, not requests — Cerebras' free tier is 30K
    tok/min and a condensed payload is ~5K tok, so 2 parallel calls is the safe
    speedup. Local ollama queues excess requests itself.
    """
    t0 = time.monotonic()
    print(f"STAGE 3 — extract for {period.human()}  [backend={backend}]\n")
    counters = {"calls": 0, "rows": 0, "err": 0, "chars": 0,
                "prompts": 0, "det": 0, "det_rows": 0}
    cfg = {i["slug"]: i for i in load_sources(include_disabled=True)}
    day = SNAPDIR / period.label
    if not day.exists():
        sys.exit(f"no snapshots at {day} — run harvest first")
    pdir = day / "prompts"
    pdir.mkdir(parents=True, exist_ok=True)
    backends = {"ollama": call_ollama, "mimo": call_mimo,
                "cerebras": call_cerebras}

    def work(meta_path: pathlib.Path) -> tuple[str, dict]:
        """One snapshot end-to-end. Every file it touches is per-stem, so many
        of these can run at once without touching shared state."""
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        stem = f"{meta['slug']}__{meta['part']}"
        local = dict.fromkeys(counters, 0)

        if meta["status"] != "FETCHED":
            write_result(day, stem, {"status": meta["status"], "rows": []})
            return f"  {stem:<24} skipped ({meta['status']})", local

        raw_html = (day / f"{stem}.html").read_text(encoding="utf-8", errors="replace")

        # LADDER STEP 1 — deterministic. If the venue publishes machine-readable
        # timestamps, parse them and skip the model entirely: free, instant, and
        # incapable of inventing a show that isn't there.
        want = cfg.get(meta["slug"], {}).get("extract", "auto")
        if want != "llm":
            strat, rows = deterministic_rows(raw_html, period, want)
            if rows:
                write_result(day, stem, {"status": "OK", "published_through": "",
                                         "method": strat, "rows": rows})
                local["det"] += 1
                local["det_rows"] += len(rows)
                return (f"  {stem:<24} {'OK':<16} rows={len(rows):<3} "
                        f"{'—':>8}  via {strat} (no model call)"), local

        # LADDER STEP 2 — the model, on condensed prose the parsers can't read.
        tp = day / f"{stem}.txt"
        text = tp.read_text(encoding="utf-8") if tp.exists() else condense(raw_html, period)
        prompt = EXTRACT_PROMPT.format(
            name=meta["name"], ts=meta["fetched_at"], url=meta["url"],
            start=period.start.isoformat(), end=period.end.isoformat(), html=text)
        (pdir / f"{stem}.txt").write_text(prompt, encoding="utf-8")

        local["chars"] += len(prompt)
        local["prompts"] += 1
        if backend == "none":
            return f"  {stem:<24} prompt written ({len(prompt)} chars)", local

        c0 = time.monotonic()
        try:
            raw = backends[backend](prompt, model)
            result = parse_json_loose(raw)
            result.setdefault("status", "OK")
            result.setdefault("rows", [])
        except Exception as exc:
            # A bad page is a localised, diagnosable failure — not a dead run.
            result = {"status": "EXTRACT_FAILED", "rows": [], "error": str(exc)[:300]}
        call_s = time.monotonic() - c0
        local["calls"] += 1
        local["rows"] += len(result.get("rows", []))
        local["err"] += result["status"] == "EXTRACT_FAILED"
        write_result(day, stem, result)
        return (f"  {stem:<24} {result['status']:<16} rows={len(result.get('rows', [])):<3} "
                f"{fmt_dur(call_s):>8}  {result.get('published_through', '')}"
                f"{result.get('error', '')}"), local

    metas = sorted(day.glob("*.meta.json"))
    if backend == "none":
        # No model calls, so no latency to hide: sequential keeps the log stable.
        for line, local in (work(mp) for mp in metas):
            print(line)
            for k in counters:
                counters[k] += local[k]
    else:
        with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
            futs = [ex.submit(work, mp) for mp in metas]
            for fut in as_completed(futs):
                line, local = fut.result()
                print(line)
                for k in counters:
                    counters[k] += local[k]

    n_calls, n_rows = counters["calls"], counters["rows"]
    n_err, sent_chars = counters["err"], counters["chars"]
    n_prompts, n_det, n_det_rows = (counters["prompts"], counters["det"],
                                    counters["det_rows"])

    if backend == "none":
        print(f"\nprompts -> {pdir}\n"
              "Paste each into a chat model, save the JSON reply as "
              "snapshots/<period>/<stem>.result.json, then run `verify`.")
        stage_done("extract", t0,
                   f"{n_det} venue(s) parsed deterministically ({n_det_rows} rows, no model) · "
                   f"{n_prompts} prompt(s) written · ~{sent_chars/3.5:,.0f} tok payload")
    else:
        el = max(time.monotonic() - t0, 1e-6)
        stage_done("extract", t0,
                   f"{n_det} deterministic ({n_det_rows} rows, no model) · "
                   f"{n_calls} model call(s) via {backend}/{model} ({n_rows} rows) · "
                   f"{n_err} error(s) · ~{sent_chars/3.5:,.0f} tok in · "
                   f"{fmt_dur(el/max(n_calls,1))}/call")
    return 0


def write_result(day: pathlib.Path, stem: str, payload: dict) -> None:
    (day / f"{stem}.result.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# ================================================================= stage 4
def verify(period: Period) -> int:
    """Two mechanical gates per row, no model involved:

       1. citat_sursa must literally occur in the snapshot  -> anti-hallucination
       2. data must fall inside the requested period        -> anti-drift

    This is 'verifica fiecare rand in sursa oficiala' as an assertion rather than
    an instruction, so it cannot be skipped or falsely claimed.
    """
    t0 = time.monotonic()
    print(f"STAGE 4 — verify for {period.human()}\n")
    day = SNAPDIR / period.label
    if not day.exists():
        sys.exit(f"no snapshots at {day} — run harvest first")
    verified, dropped, n_badlink = [], [], 0

    for res_path in sorted(day.glob("*.result.json")):
        stem = res_path.name.removesuffix(".result.json")
        result = json.loads(res_path.read_text(encoding="utf-8"))
        meta = json.loads((day / f"{stem}.meta.json").read_text(encoding="utf-8"))
        # Evidence = the condensed text the model actually saw, which is itself
        # derived deterministically from the snapshot. Checking against raw HTML
        # would reject honest quotes that span tags ("Sala Mică 19:30").
        ev = day / f"{stem}.txt"
        if not ev.exists():
            ev = day / f"{stem}.html"
        norm = WS_RE.sub(" ", ev.read_text(encoding="utf-8", errors="replace")) if ev.exists() else ""
        # Links are checked against the FULL page, not the condensed text. A
        # deterministic extractor reads hrefs straight from the markup, and those
        # lines rarely sit near a date match, so condensing drops them — checking a
        # real link against the condensed view flagged 55 genuine TNB URLs as
        # invented. The right evidence for a URL is the page it came from.
        hp_full = day / f"{stem}.html"
        link_ev = (WS_RE.sub(" ", hp_full.read_text(encoding="utf-8", errors="replace"))
                   if hp_full.exists() else norm)

        kept = drop = 0
        for row in result.get("rows", []):
            row |= {"slug": meta["slug"], "institutie": meta["name"],
                    "sursa_url": meta["url"], "part": meta["part"]}
            quote = WS_RE.sub(" ", str(row.get("citat_sursa", ""))).strip()

            if not quote or quote not in norm:
                row["motiv"] = "citat_sursa not found in snapshot — possible hallucination"
                dropped.append(row); drop += 1
                continue
            try:
                d = dt.date.fromisoformat(str(row.get("data", period.start.isoformat())))
            except ValueError:
                row["motiv"] = f"unparseable data field: {row.get('data')!r}"
                dropped.append(row); drop += 1
                continue
            if not period.contains(d):
                row["motiv"] = f"date {d} outside requested period {period.label}"
                dropped.append(row); drop += 1
                continue

            # A link the model produced that does not occur in the evidence is
            # invented. Don't drop an otherwise-good row for it — fall back to the
            # institution's page and record that we did.
            link = str(row.get("link") or "").strip()
            if link and link not in link_ev and link not in norm:
                row["link_invented"] = link
                row["link"] = meta["url"]
                n_badlink += 1

            row["data"] = d.isoformat()
            verified.append(row); kept += 1

        print(f"  {stem:<24} {result.get('status','?'):<16} kept {kept} dropped {drop}")

    # Same performance can surface in overlapping parts; collapse exact duplicates.
    seen, unique = set(), []
    for r in verified:
        key = (r["slug"], r.get("data"), r.get("ora"), r.get("spectacol"))
        if key not in seen:
            seen.add(key)
            unique.append(r)
    deduped = len(verified) - len(unique)

    OUTDIR.mkdir(exist_ok=True)
    out = OUTDIR / f"{period.label}.verified.json"
    out.write_text(json.dumps({"period": period.label, "verified": unique,
                               "dropped": dropped}, ensure_ascii=False, indent=2),
                   encoding="utf-8")
    print(f"\n-> {out}")
    stage_done("verify", t0,
               f"{len(unique)} row(s) verified · {len(dropped)} dropped · "
               f"{deduped} duplicate(s) collapsed"
               + (f" · {n_badlink} invented link(s) replaced" if n_badlink else ""))
    return 0


# ================================================================= stage 5
def _cell(value) -> str:
    """Single-line display form of a cell value (titles do contain newlines)."""
    return re.sub(r"\s+", " ", str(value)).strip() or "—"


def _print_table(header: list[str], rows: list[list[str]]) -> None:
    """Console view of the same data that goes into the spreadsheet."""
    cols = list(zip(header, *rows)) if rows else [[h] for h in header]
    widths = [max(len(c) for c in col) for col in cols]
    print("  ".join(h.ljust(w) for h, w in zip(header, widths)))
    for r in rows:
        print("  ".join(c.ljust(w) for c, w in zip(r, widths)))


def _write_xlsx(path: pathlib.Path, header, rows, status_header, status_rows) -> None:
    """Primary output: a real .xlsx (openpyxl) — events sheet + status sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    def fill(ws, head, data):
        ws.append(head)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in data:
            ws.append(r)
        cols = list(zip(head, *data)) if data else [[h] for h in head]
        for i, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = \
                min(max(len(str(c)) for c in col) + 2, 60)
        ws.freeze_panes = "A2"

    wb = Workbook()
    ws = wb.active
    ws.title = "Spectacole"
    fill(ws, header, rows)
    fill(wb.create_sheet("Stare verificare"), status_header, status_rows)
    wb.save(path)


def _write_csv(path: pathlib.Path, header, rows) -> None:
    """Fallback output: flat .csv (events only — a csv has no second sheet).
    utf-8-sig so Excel reads the Romanian diacritics without a manual import.
    """
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


def render(period: Period) -> int:
    """Always emit the event list (.xlsx, .csv if openpyxl is missing).
    Partial results are first-class; failure is a status."""
    t0 = time.monotonic()
    day = SNAPDIR / period.label
    vpath = OUTDIR / f"{period.label}.verified.json"
    verified = json.loads(vpath.read_text(encoding="utf-8"))["verified"] if vpath.exists() else []

    verified.sort(key=lambda r: (r.get("data", ""), r.get("ora", "99:99")))
    header = ([] if period.single_day else ["Data"]) + \
             ["Ora", "Instituția", "Spectacol", "Sala", "Link de verificare"]
    rows: list[list[str]] = []
    for row in verified:
        cells = [_cell(row.get("ora")), _cell(row["institutie"]),
                 _cell(row.get("spectacol")), _cell(row.get("sala")),
                 row.get("link") or row["sursa_url"]]
        if not period.single_day:
            cells.insert(0, dt.date.fromisoformat(row["data"]).strftime("%d.%m"))
        rows.append(cells)

    # Per-institution verification status — second sheet, so the workbook
    # answers "which venues did we actually check?" next to the events.
    status_header = ["Instituția", "Stare", "Sursă", "Verificat la"]
    status_rows: list[list[str]] = []
    for inst in load_sources(include_disabled=True):
        slug = inst["slug"]
        if not inst.get("enabled", True):
            status_rows.append([inst["name"],
                                f"Exclus — {inst.get('disabled_reason', 'dezactivat')}",
                                "—", "—"])
            continue
        metas = [json.loads(p.read_text(encoding="utf-8"))
                 for p in sorted(day.glob(f"{slug}__*.meta.json"))]
        results = [json.loads(p.read_text(encoding="utf-8"))
                   for p in sorted(day.glob(f"{slug}__*.result.json"))]
        n = sum(1 for r in verified if r["slug"] == slug)
        codes = [r.get("status") for r in results] + [m.get("status") for m in metas]
        code = next((c for c in codes if c not in (None, "FETCHED")), "FETCHED")
        state = (f"{n} spectacol(e)" if n else
                 {"NOT_PUBLISHED": "Nu este publicat încă",
                  "NO_SCHEDULE_FOUND": "Calendar negăsit",
                  "NEEDS_BROWSER": "Necesită browser",
                  "EXTRACT_FAILED": "Eroare la extragere",
                  "FETCH_FAILED": "Eroare la accesare",
                  "SKIPPED_NO_BACKEND": "Extragere neexecutată",
                  # 'OK with zero kept rows' is a real outcome — the page parsed
                  # and simply had nothing in period, or every row failed verify.
                  # It used to fall through to NEVERIFICAT, which reads as "we
                  # never looked" and is exactly the ambiguity this table exists
                  # to remove.
                  "OK": "0 spectacole verificate",
                  "FETCHED": "0 spectacole în perioadă"}.get(code, "NEVERIFICAT"))
        if any(m.get("tls_verified") is False for m in metas):
            state += " ⚠︎ TLS neverificat"
        src = metas[0]["url"] if metas else "—"
        when = metas[0]["fetched_at"][:19].replace("T", " ") if metas else "—"
        status_rows.append([inst["name"], state, src, when])

    OUTDIR.mkdir(exist_ok=True)
    try:
        import openpyxl  # noqa: F401  (availability check — primary format)
        out = OUTDIR / f"{period.label}.xlsx"
        _write_xlsx(out, header, rows, status_header, status_rows)
    except ImportError:
        out = OUTDIR / f"{period.label}.csv"
        _write_csv(out, header, rows)
        print("  ⚠︎  openpyxl missing — wrote .csv fallback "
              "(pip install openpyxl for .xlsx)")

    if rows:
        _print_table(header, rows)
    else:
        print("  " + "  ".join(header))
        print("  (niciun spectacol verificat)")
    print()
    _print_table(status_header, status_rows)
    print(f"\n-> {out}")
    venues = len({r["slug"] for r in verified})
    stage_done("render", t0,
               f"{len(verified)} performance(s) across {venues} venue(s) -> {out.name}")
    return 0


# ================================================================= cli
def main() -> int:
    ap = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="periods: current-month | next-month | today | YYYY-MM | "
               "YYYY-MM-DD | YYYY-MM-DD..YYYY-MM-DD")
    ap.add_argument("stage", choices=["preflight", "harvest", "discover", "diagnose",
                                      "extract", "verify", "render", "all"])
    ap.add_argument("--period", default=None, help="see epilog; default current-month")
    ap.add_argument("--date", default=None, help="alias for --period <single day>")
    ap.add_argument("--backend", default=os.environ.get("TEATRE_BACKEND", "none"),
                    choices=["none", "ollama", "cerebras", "mimo"],
                    help="LLM backend for stage 3 (default: none = write prompts only)")
    ap.add_argument("--model", default=os.environ.get("TEATRE_MODEL", ""),
                    help="model name, e.g. gpt-oss-120b or xiaomi/mimo-v2.5")
    ap.add_argument("--concurrency", type=int,
                    default=int(os.environ.get("TEATRE_CONCURRENCY", "4")),
                    help="parallel workers for preflight/harvest network calls "
                         "(default 4, env TEATRE_CONCURRENCY)")
    ap.add_argument("--extract-concurrency", type=int,
                    default=int(os.environ.get("TEATRE_EXTRACT_CONCURRENCY", "2")),
                    help="parallel model calls in extract; keep low — API tiers "
                         "limit tokens/minute, not requests (default 2, env "
                         "TEATRE_EXTRACT_CONCURRENCY)")
    args = ap.parse_args()

    try:
        period = parse_period(args.period or args.date or "current-month")
    except ValueError as exc:
        sys.exit(f"bad period {args.period or args.date!r}: {exc}")

    model = args.model or {"ollama": "qwen3-coder:latest",
                           "cerebras": "gpt-oss-120b",
                           "mimo": "xiaomi/mimo-v2.5"}.get(args.backend, "")

    if args.stage == "preflight":
        return preflight(concurrency=args.concurrency)[0]
    if args.stage == "extract":
        return extract(period, args.backend, model, args.extract_concurrency)
    if args.stage == "all":
        code, _ = preflight(concurrency=args.concurrency)
        if code:
            print("\nABORTED: systemic failure, not a site problem. "
                  "Fix the environment; do not guess at causes.")
            return 1
        print()
        harvest(period, args.concurrency)
        print()
        diagnose(period)
        print()
        extract(period, args.backend, model, args.extract_concurrency)
        print()
        verify(period)
        print()
        code = render(period)
        print_rollup()
        return code
    if args.stage == "harvest":
        return harvest(period, args.concurrency)
    return {"discover": discover, "diagnose": diagnose,
            "verify": verify, "render": render}[args.stage](period)


if __name__ == "__main__":
    raise SystemExit(main())
